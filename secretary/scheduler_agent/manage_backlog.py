from datetime import (datetime, timedelta)
from constants import (MONGO_DEFAULT_DB, MONGO_HOST, MONGO_PASS, MONGO_PORT, MONGO_USER)
from faker import Faker
from find_manage import *
from interfaces import *
from insert_manage import *
from fake_profile import *
from manage_visits import *
import random
from openpyxl import load_workbook
import logging

fake = Faker('pt_BR')
logger = logging.getLogger()
logger.setLevel(logging.DEBUG)

class excel:
    def __init__(self):
        super(excel, self).__init__()
        self.ws = []

    def generate_timetable_with_backlog(self, start_create: datetime, finished_create:datetime):
        """Inicializa variaveis necessarias e lista todos os customers sem visita e cria visitas."""
        customer_no_have_date, customer_have_date = excel.customer_available()
        for customer in customer_have_date:
            supplier = Getter.get_supplier_in_backlog('customer', customer)
            supplier = supplier[0]
            start_date = Getter.find_start_date_in_backlog('customer', customer)
            start_date = start_date[0]
            start_date = datetime.strptime(start_date, "%Y-%m-%dT%H")
            start_date = Manage.available_date(start_date)
            valida = excel.validate_date_avaible_for_supplier(supplier, start_date) 
            while valida is None:
                start_date = Manage.date_sum_hour(start_date, 1)
                start_date = Manage.available_date(start_date)
                start_date = Manage.available_date(start_date)
                valida = excel.validate_date_avaible_for_supplier(supplier, start_date)
            """atualiza as condições para gerar as visitas ou não"""
            """start_date = finished_create o sistema para de gerar visitas."""
            condition_finish_create_visits = start_date >= finished_create
            conditon_continue_create_visits = start_date < finished_create
            if conditon_continue_create_visits:
                end_date = Manage.date_sum_hour(start_date, 1)
                payload = Manage.generate_available_payload_visit(customer, supplier, 'Backlog', 'Instalação de Modem', start_date, end_date)
                Manage.insert_payload(payload)
        start_date = start_create
        start_date = Manage.available_date(start_date)
        for customer in customer_no_have_date:
            supplier = Getter.get_supplier_in_backlog('customer', customer)
            supplier = supplier[0]
            valida = excel.validate_date_avaible_for_supplier(supplier, start_date) 
            while valida is None:
                start_date = Manage.date_sum_hour(start_date, 1)
                start_date = Manage.available_date(start_date)
                valida = excel.validate_date_avaible_for_supplier(supplier, start_date)
            """atualiza as condições para gerar as visitas ou não"""
            """start_date = finished_create o sistema para de gerar visitas."""
            condition_finish_create_visits = start_date >= finished_create
            conditon_continue_create_visits = start_date < finished_create
            if conditon_continue_create_visits:
                end_date = Manage.date_sum_hour(start_date, 1)
                payload = Manage.generate_available_payload_visit(customer, supplier, 'Pendente', 'Instalação de Modem', start_date, end_date)
                Manage.insert_payload(payload)            
        return True

    def customer_available(self):
        try:
            customer_backlog = Getter.get_all_backlog('customer')
            customer_no_have_date = Getter.find_all_customer_have_date('backlog', None)
            customer_have_date = [x for x in customer_backlog if x not in customer_no_have_date]
            return customer_no_have_date, customer_have_date
        except:
            return 'falha'

    def insert_backlog_in_db(self):
        try:
            excel.load_backlog()
            row, column = excel.max_row_column()
            for r in range(1, row+1):
                customer = excel.read_cell(r, 1)
                supplier = excel.read_cell(r, 2)
                task = excel.read_cell(r, 3)
                start_date = excel.read_cell(r, 4)
                if start_date is None:
                    payload = Manage.generate_none_payload_visit(customer, supplier, task)
                else:
                    end_date = Manage.date_sum_hour(start_date, 1)
                    payload = Manage.generate_available_payload_visit(customer, supplier, 'Backlog', 'Instalação de Modem', start_date, end_date)
                ''' inserir data in mongo '''
                f = Insert_Backlog_Payload()
                f.generate(payload)
                f.insert_to_mongo()
            return True
        except:
            return False

    def load_backlog(self):
        """ carrega um backlog em self.ws """
        try:
            print('insira o caminho para o arquivo de backlog: ')
            diretorio = input('')
            wb = load_workbook(diretorio)
            self.ws = wb.active
            return self.ws
        except Exception as log:
            print('Falha ao encontrar backlog')
            return None
    
    def validate_date_avaible_for_supplier(self, supplier, start_date):
        '''retorna True se caso o supplier estiver disponivel e False caso contrario'''
        try:
            list_suppliers = Getter.find_all_suppliers_in_this_date('time_table', start_date.strftime("%Y-%m-%dT%H"))
            for suppliers in list_suppliers:
                if supplier == suppliers:
                    return None
            return True    
        except:
            print("falha ao validar data")

    def read_cell(self, nrow, ncolumn):
        '''ler celula e retornar data'''
        try:
            data = self.ws.cell(nrow, ncolumn).value
            return data
        except:
            
            return None

    def max_row_column(self):
        """retorna o numero maximo de linhas e colunas"""
        try:
            ws = self.ws
            rows = ws.max_row
            columns = ws.max_column
            return rows, columns
        except:
            print("não foi possivel calcular os valores maximos")
            return None, None


if __name__ == '__main__':
    excel = excel()
    excel.insert_backlog_in_db()
    start_create = fake.future_datetime("+10h")
    finished_create = Manage.date_sum_hour(start_create, 19)
    print("A visitas do log serão geradas de {} até {}".format(start_create.strftime("%Y-%m-%dT%H"), finished_create.strftime("%Y-%m-%dT%H")))
    excel.generate_timetable_with_backlog(start_create, finished_create)